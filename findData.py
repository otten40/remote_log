import re
import os
import openpyxl
from tkinter import *


# makes the location of the spreadsheet the active directory and opens the remote log spreadsheet
os.chdir('C:\\Users\\320051\\Desktop')
remote_log = openpyxl.load_workbook('Updated Remote Access Log copy.xlsx')
sheet = remote_log['Sheet1']
# --------------------------------------------------------------------------------------------------


# -------------------------------------------------------------
# GUI code
def get_data(event=None):
    return str_var.get()
    global root
    root.quit()


root = Tk()

str_var = StringVar()

str_var.set("Paste Email")

str_entry = Entry(root, textvariable=str_var)
str_entry.pack(side=LEFT)

get_data_button = Button(root, text="Submit")
get_data_button.bind("<Button-1>", get_data)
get_data_button.pack(side=LEFT)

root.mainloop()
# ----------------------------------------------------------------


# This code takes the pasted body of the email and turns it into a string that can be multiple lines and assigns it
# to the body variable
lines = []
while True:
    line = get_data()
    if line:
        lines.append(line)
    else:
        break

body = '\n'.join(lines)


# This code finds the date in the body of the email
def find_date(email_body):
    req_date_regex = re.compile(r'(January|February|March|April|May|June|July|August|September|October|November|'
                                r'December) (\d\d), (\d\d\d\d)')
    req_date = req_date_regex.search(email_body)
    month = req_date.group(1)
    day = req_date.group(2)
    year = req_date.group(3)
    date = month + ' ' + day + ',' + ' ' + year
    print(date)
    return date


# This code finds the time in in the body of the email
def find_time_in(email_body):
    req_time_regex = re.compile(r'(\d\d:\d\d):(\d\d) (AM|PM)')
    req_time = req_time_regex.search(email_body)
    req_time = req_time.group(1) + ' ' + req_time.group(3)
    print(req_time)
    return req_time


find_time_in(body)


# This code finds the user name in the body of the email
def find_name(email_body):
    req_name_regex = re.compile('Requesting User: (.+?)\\n')
    req_name = req_name_regex.search(email_body)
    req_name = req_name.group(1)
    print(req_name)
    return req_name


find_name(body)


# This code finds the reason for remote access in the body of the email
def find_reason(email_body):
    reason_regex = re.compile('Reason:\\n(.+)')
    reason = reason_regex.search(email_body)
    reason = reason.group(1)
    print(reason)
    return reason


find_reason(body)


# This code finds the reason for request number in the body of the email
def find_req_num(email_body):
    req_num_regex = re.compile('Request # (.+?)\\n')
    req_num = req_num_regex.search(email_body)
    req_num = req_num.group(1)
    print(req_num)
    return req_num


find_req_num(body)


# This code finds the ticket number in the body of the email
def find_tick_num(email_body):
    tick_num_regex = re.compile(r'((\d-\d\d\d\d\d\d\d\d\d\d)*)')
    tick_num = tick_num_regex.search(email_body)
    tick_num = tick_num.group(1)
    print(tick_num)
    return tick_num


remote_log_fields = {'date': find_date(get_data()),
                     'time_in': find_time_in(get_data()),
                     'user': find_name(get_data()),
                     'reason': find_reason(get_data()),
                     'ticket': find_tick_num(get_data()),
                     'req_num': find_req_num(get_data()),
                     # 'it_rep': find_tech(body),
                     # 'badge_num': find_badge_num(body),
                     # 'time_out': find_time_out(body)
                     # 'company': find_company,
                     }


sheet.cell(row=sheet.max_row + 1, column=1).value = remote_log_fields['date']
sheet.cell(row=sheet.max_row, column=2).value = remote_log_fields['time_in']
sheet.cell(row=sheet.max_row, column=3).value = remote_log_fields['user']
sheet.cell(row=sheet.max_row, column=5).value = remote_log_fields['reason']
sheet.cell(row=sheet.max_row, column=6).value = remote_log_fields['ticket']
sheet.cell(row=sheet.max_row, column=7).value = remote_log_fields['req_num']


remote_log.save('updated Remote Access Log copy.xlsx')
